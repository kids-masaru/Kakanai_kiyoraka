
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from pathlib import Path
from typing import Optional, Tuple
import os

CONFIG_DIR = Path(__file__).parent.parent / "config"
TEMPLATE_PATH = CONFIG_DIR / "template_csv.xlsx"

class CSVService:
    def convert_csv_to_excel(self, file_content: bytes, original_filename: str) -> Tuple[Optional[bytes], Optional[str]]:
        """
        CSVバイナリデータを受け取り、Excelファイル(bytes)とファイル名を返す
        エラー時は (None, error_message) を返す
        """
        # テンプレート確認
        if not TEMPLATE_PATH.exists():
            return None, "Template file not found on server."

        # CSV読み込み
        encodings = ['utf-8-sig', 'utf-8', 'shift-jis', 'cp932', 'iso-2022-jp', 'euc-jp']
        df = None
        last_error = None
        
        for encoding in encodings:
            try:
                # バイト列を読み込む
                df = pd.read_csv(
                    BytesIO(file_content), 
                    encoding=encoding, 
                    header=0, 
                    on_bad_lines='warn',
                    index_col=False
                )
                break
            except Exception as e:
                last_error = str(e)
                continue
        
        if df is None:
            return None, f"Failed to read CSV. Please check encoding. Last error: {last_error}"

        try:
            # Excel生成
            wb = openpyxl.load_workbook(TEMPLATE_PATH)
            
            sheet_name = "貼り付け用"
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # クリア
                for row in ws.iter_rows():
                    for cell in row:
                        cell.value = None
            else:
                ws = wb.create_sheet(sheet_name)
            
            # 書き込み (Indexなし、Headerあり)
            rows = dataframe_to_rows(df, index=False, header=True)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            # バイト列に出力
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # ファイル名生成
            base_name = os.path.splitext(original_filename)[0]
            new_filename = f"{base_name}_converted.xlsx"
            
            return output.getvalue(), new_filename
            
        except Exception as e:
            return None, f"Excel conversion failed: {str(e)}"

csv_service = CSVService()
